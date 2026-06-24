"""
Generate an Excel workbook that computes pole/post embedment depth per
IBC 2021 Section 1807.3, with SI inputs/outputs and live formulas.

Run:  python make_xlsx.py
Output: Pole_Embedment_IBC2021_SI.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ---------- styling helpers ----------
NAVY = "0F172A"
BRAND = "0E7490"
LIGHT = "ECFEFF"
HEAD = "155E75"
GREY = "F1F5F9"
LINE = "CBD5E1"

f_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
f_sub = Font(name="Calibri", size=10, color="DBEAFE")
f_h = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
f_lbl = Font(name="Calibri", size=10, bold=True)
f_in = Font(name="Calibri", size=11, bold=True, color="155E75")
f_unit = Font(name="Calibri", size=9, color="64748B")
f_note = Font(name="Calibri", size=9, italic=True, color="64748B")
f_res = Font(name="Calibri", size=18, bold=True, color="155E75")
f_resl = Font(name="Calibri", size=9, color="64748B")

fill_title = PatternFill("solid", fgColor=NAVY)
fill_head = PatternFill("solid", fgColor=HEAD)
fill_in = PatternFill("solid", fgColor="FFFFFF")
fill_res = PatternFill("solid", fgColor=LIGHT)
fill_grey = PatternFill("solid", fgColor=GREY)
in_border = Border(*(Side(style="thin", color="94A3B8"),) * 4)
thin = Side(style="thin", color=LINE)
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = "Calculator"
ws.sheet_view.showGridLines = False

widths = {"A": 2.5, "B": 30, "C": 16, "D": 13, "E": 14, "F": 14, "G": 14, "H": 15}
for c, w in widths.items():
    ws.column_dimensions[c].width = w


def merge(rng, value, font=None, fill=None, align="left", border=None):
    ws.merge_cells(rng)
    first = rng.split(":")[0]
    cell = ws[first]
    cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if border:
        for row in ws[rng]:
            for cc in row:
                cc.border = border
    return cell


# ---------- Title ----------
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 16
merge("B1:H1", "Pole / Post Embedment Depth  —  IBC 2021 §1807.3 (SI units)",
      font=f_title, fill=fill_title, align="left")
merge("B2:H2", "Lateral-load embedment for round/square posts. Inputs in SI; code equations applied internally.",
      font=f_sub, fill=fill_title, align="left")

# ---------- INPUTS ----------
merge("B4:H4", "INPUTS", font=f_h, fill=fill_head)

inputs = [
    ("Applied lateral force, P", 10, "kN", "Horizontal design force."),
    ("Load height above grade, h", 2.5, "m", "Ground surface to point of application of P."),
    ("Post width / diameter, b", 0.30, "m", "Round dia. or diagonal of square post/footing."),
    ("Allowable lateral bearing, S", 23.6, "kPa/m", "Per m of depth (= kN/m^3), Table 1806.2."),
    ("Isolated-pole factor (1 or 2)", 1, "—", "2 = double bearing per §1806.3.4."),
    ("Apply 12 ft (3.66 m) depth cap (1=yes,0=no)", 1, "—", "Freeze bearing at code depth limit."),
]
r = 5
input_cells = {}
for name, val, unit, note in inputs:
    ws[f"B{r}"] = name
    ws[f"B{r}"].font = f_lbl
    ws[f"B{r}"].alignment = Alignment(vertical="center", wrap_text=True)
    c = ws[f"C{r}"]
    c.value = val
    c.font = f_in
    c.fill = fill_in
    c.border = in_border
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws[f"D{r}"] = unit
    ws[f"D{r}"].font = f_unit
    ws[f"D{r}"].alignment = Alignment(horizontal="left", vertical="center")
    merge(f"E{r}:H{r}", note, font=f_note, align="left")
    input_cells[name] = f"C{r}"
    r += 1

P = input_cells["Applied lateral force, P"]
H = input_cells["Load height above grade, h"]
B = input_cells["Post width / diameter, b"]
S = input_cells["Allowable lateral bearing, S"]
ISO = input_cells["Isolated-pole factor (1 or 2)"]
CAP = input_cells["Apply 12 ft (3.66 m) depth cap (1=yes,0=no)"]

# Soil reference note
r += 0
merge(f"B{r}:H{r}",
      "Table 1806.2 (kPa/m): Bedrock 188.5 | Rock 62.8 | Sandy gravel GW,GP 31.4 | "
      "Sand SW,SP,SM,SC,GM,GC 23.6 | Clay/silt CL,ML,MH,CH 15.7",
      font=f_note, align="left")
r += 2

# ---------- UNIT CONVERSION (internal) ----------
conv_start = r
merge(f"B{r}:H{r}", "INTERNAL CONVERSION TO US-CUSTOMARY", font=f_h, fill=fill_head)
r += 1
conv = [
    ("P  (lb)", f"={P}*224.808943"),
    ("h  (ft)", f"={H}*3.280839895"),
    ("b  (ft)", f"={B}*3.280839895"),
    ("Effective lateral bearing s  (psf/ft)", f"={S}*6.36588*{ISO}"),
]
conv_cells = {}
for name, formula in conv:
    ws[f"B{r}"] = name
    ws[f"B{r}"].font = f_lbl
    c = ws[f"C{r}"]
    c.value = formula
    c.fill = fill_grey
    c.border = box
    c.alignment = Alignment(horizontal="center")
    c.number_format = "0.000"
    conv_cells[name] = f"C{r}"
    r += 1

Plb = conv_cells["P  (lb)"]
Hft = conv_cells["h  (ft)"]
Bft = conv_cells["b  (ft)"]
s = conv_cells["Effective lateral bearing s  (psf/ft)"]
r += 1

# ---------- CONSTRAINED ----------
merge(f"B{r}:H{r}", "CONSTRAINED  (restrained at grade — IBC Eq. 18-3)", font=f_h, fill=fill_head)
r += 1
# d uncapped = cbrt(4.25 P h /(s b)); if capped & >12 -> S3 frozen at 12ft
ws[f"B{r}"] = "d, uncapped (ft)"
ws[f"B{r}"].font = f_lbl
ws[f"C{r}"] = f"=(4.25*{Plb}*{Hft}/({s}*{Bft}))^(1/3)"
con_unc = f"C{r}"
ws[con_unc].number_format = "0.000"; ws[con_unc].fill = fill_grey; ws[con_unc].border = box
ws[con_unc].alignment = Alignment(horizontal="center")
r += 1
ws[f"B{r}"] = "d, final (ft)"
ws[f"B{r}"].font = f_lbl
ws[f"C{r}"] = (f"=IF(AND({CAP}=1,{con_unc}>12),"
               f"SQRT(4.25*{Plb}*{Hft}/({s}*12*{Bft})),{con_unc})")
con_ft = f"C{r}"
ws[con_ft].number_format = "0.000"; ws[con_ft].fill = fill_grey; ws[con_ft].border = box
ws[con_ft].alignment = Alignment(horizontal="center")
r += 1

# ---------- NON-CONSTRAINED iteration table ----------
merge(f"B{r}:H{r}", "NON-CONSTRAINED  (free top — IBC Eq. 18-1 & 18-2, solved by iteration)",
      font=f_h, fill=fill_head)
r += 1
hdr_row = r
headers = ["Iter n", "d guess (ft)", "depth for S (ft)", "S1 (psf)", "A (ft)", "d new (ft)"]
for i, htxt in enumerate(headers):
    col = get_column_letter(3 + i)  # C..H
    cell = ws[f"{col}{hdr_row}"]
    cell.value = htxt
    cell.font = Font(bold=True, size=9, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=BRAND)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = box
ws[f"B{hdr_row}"] = "iteration"
ws[f"B{hdr_row}"].font = f_note
r += 1

N_ITER = 40
first_iter = r
for k in range(N_ITER):
    rr = r + k
    cN, cdg, cdfs, cS1, cA, cdn = (f"C{rr}", f"D{rr}", f"E{rr}", f"F{rr}", f"G{rr}", f"H{rr}")
    ws[cN] = k
    if k == 0:
        ws[cdg] = f"={Hft}"          # initial guess = h (ft)
    else:
        ws[cdg] = f"=H{rr-1}"        # previous d_new
    ws[cdfs] = f"=IF({CAP}=1,MIN({cdg},12),{cdg})"
    ws[cS1] = f"={s}*{cdfs}/3"
    ws[cA] = f"=2.34*{Plb}/({cS1}*{Bft})"
    ws[cdn] = f"=0.5*{cA}*(1+SQRT(1+4.36*{Hft}/{cA}))"
    for cc in (cN, cdg, cdfs, cS1, cA, cdn):
        ws[cc].alignment = Alignment(horizontal="center")
        ws[cc].number_format = "0.0000" if cc != cN else "0"
        ws[cc].border = box
        ws[cc].font = Font(size=9)
last_iter = r + N_ITER - 1
non_ft = f"H{last_iter}"   # converged d_new (ft)
r = last_iter + 2

# ---------- RESULTS (placed prominently) ----------
res_row = r
merge(f"B{r}:H{r}", "RESULTS — Required embedment depth, d", font=f_h, fill=fill_head)
r += 1
ws.row_dimensions[r].height = 40
# Non-constrained result
merge(f"B{r}:C{r}", "Non-constrained, d", font=f_lbl, fill=fill_res, align="center", border=box)
nc = ws[f"D{r}"]
nc.value = f"={non_ft}*0.3048"
nc.font = f_res; nc.fill = fill_res; nc.border = box
nc.number_format = "0.000"
nc.alignment = Alignment(horizontal="center", vertical="center")
merge(f"E{r}:E{r}", "m", font=f_resl, fill=fill_res, align="left", border=box)
# Constrained result
merge(f"F{r}:F{r}", "Constr., d", font=f_lbl, fill=fill_res, align="center", border=box)
cc2 = ws[f"G{r}"]
cc2.value = f"={con_ft}*0.3048"
cc2.font = f_res; cc2.fill = fill_res; cc2.border = box
cc2.number_format = "0.000"
cc2.alignment = Alignment(horizontal="center", vertical="center")
merge(f"H{r}:H{r}", "m", font=f_resl, fill=fill_res, align="left", border=box)
r += 2

# ---------- Notes / formulas ----------
merge(f"B{r}:H{r}", "METHOD & FORMULAS", font=f_h, fill=fill_head)
r += 1
notes = [
    "Non-constrained:  A = 2.34·P/(S1·b) ;  d = 0.5·A·(1 + SQRT(1 + 4.36·h/A)).  S1 = lateral bearing at d/3 (iterated).",
    "Constrained:  d = SQRT(4.25·P·h/(S3·b)).  S3 = lateral bearing at full depth d  (closed form: d = (4.25·P·h/(s·b))^(1/3)).",
    "Conversions: 1 m = 3.280839895 ft ; 1 kN = 224.808943 lb ; 1 kPa/m (kN/m^3) = 6.36588 psf/ft ; d(m) = d(ft)·0.3048.",
    "Deflection basis is 12 mm (0.5 in) at grade. Depth limited to 3.66 m (12 ft) for computing lateral pressure (§1807.3).",
    "Preliminary design only — verify with governing code edition and project geotechnical data.",
]
for nt in notes:
    merge(f"B{r}:H{r}", nt, font=f_note, align="left")
    ws.row_dimensions[r].height = 26
    r += 1

# Data validation for cap/iso (lists)
dv_iso = DataValidation(type="list", formula1='"1,2"', allow_blank=False)
dv_cap = DataValidation(type="list", formula1='"0,1"', allow_blank=False)
ws.add_data_validation(dv_iso); dv_iso.add(ws[ISO])
ws.add_data_validation(dv_cap); dv_cap.add(ws[CAP])

ws.freeze_panes = "A3"

out = "Pole_Embedment_IBC2021_SI.xlsx"
wb.save(out)
print("Saved", out)
